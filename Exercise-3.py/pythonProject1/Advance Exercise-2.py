


'''Todo 1. Create a unittest for a module which has a class caled Calculator. It has 5 methods add, sub, mul, div and mod.
All these methods must be tested with positive,negative float and integers including zero.'''


# import unittest
#
#
# class calculator:
#     def add(self, a, b):
#         return a + b
#
#     def sub(self, a, b):
#         return a - b
#
#     def mul(self, a, b):
#         return a * b
#
#     def div(self, a, b):
#         return a / b
#
#     def mod(self, a, b):
#         return a % b
#
# class TestCalculator(unittest.TestCase):
#     def setUp(self):
#         self.calc = calculator()
#
#     def test_add(self):
#         self.assertEqual(self.calc.add(2, 3), 5)
#         self.assertEqual(self.calc.add(-2, -3), -5)
#         self.assertEqual(self.calc.add(2.5, 3.5), 6)
#         self.assertEqual(self.calc.add(-2.5, -3.5), -6)
#         self.assertEqual(self.calc.add(0, 0), 0)
#
#     def test_sub(self):
#         self.assertEqual(self.calc.sub(2, 3), -1)
#         self.assertEqual(self.calc.sub(-2, -3), 1)
#         self.assertEqual(self.calc.sub(2.5, 3.5), -1.0)
#         self.assertEqual(self.calc.sub(-2.5, -3.5), 1.0)
#         self.assertEqual(self.calc.sub(0, 0), 0)
#
#     def test_mul(self):
#         self.assertEqual(self.calc.mul(2, 3), 6)
#         self.assertEqual(self.calc.mul(-2, -3), 6)
#         self.assertEqual(self.calc.mul(2.5, 3.5), 8.75)
#         self.assertEqual(self.calc.mul(-2.5, -3.5), 8.75)
#         self.assertEqual(self.calc.mul(0, 0), 0)
#
#     def test_div(self):
#         self.assertEqual(self.calc.div(6, 3), 2)
#         self.assertEqual(self.calc.div(-6, -3), 2)
#         self.assertEqual(self.calc.div(6, 4), 1.5)
#         self.assertEqual(self.calc.div(-6, 4), -1.5)
#         self.assertEqual(self.calc.div(0, 5), 0)
#
#     def test_mod(self):
#         self.assertEqual(self.calc.mod(6, 4), 2)
#         self.assertEqual(self.calc.mod(-6, -4), -2)
#         self.assertEqual(self.calc.mod(6.5, 4), 2.5)
#         self.assertEqual(self.calc.mod(-6.5, 4), 1.5)
#         self.assertEqual(self.calc.mod(0, 5), 0)
#
# if __name__ == '__main__':
#     unittest.main()


'''Todo 2. Create a student report in xlsx which will have 4 columns student, maths, science, english 3 subjects marks. 
Now create a line chart for this data in xlsx.'''


import xlsxwriter

data = {
    'Student': ['Anjum', 'Bansi', 'janvi', 'Divya'],
    'Maths': [85, 90, 78, 82],
    'Science': [92, 88, 95, 84],
    'English': [80, 85, 88, 90]
}

workbook = xlsxwriter.Workbook('student_report.xlsx')       # Create a workbook and add a worksheet
worksheet = workbook.add_worksheet()


for col, column_name in enumerate(data.keys()):         # Add data to the worksheet
    worksheet.write(0, col, column_name)

    for row, value in enumerate(data[column_name], start=1):
        worksheet.write(row, col, value)

chart = workbook.add_chart({'type': 'line'})        # Create a chart object


for i in range(1, 4):               # Assuming the first column is for 'Student'
    chart.add_series({
        'name': [worksheet.name, 0, i],
        'categories': [worksheet.name, 1, 0, len(data['Student']), 0],
        'values': [worksheet.name, 1, i, len(data['Student']), i],
    })

worksheet.insert_chart('E2', chart)     # Insert the chart into the worksheet
workbook.close()                        # Close the workbook


'''Todo 3. Create a sales report in xlsx which will have the following data.
Sales Person 1 : 78,000.0
Sales Person 2 : 32,000.0
Sales Person 3 : 45,000.0
Sales Person 4 : 11,000.0
Sales Person 5 : 20,000.0
Create a Pie Chart with the above data.'''


import xlsxwriter

sales_data = {                      # Sales data
    'Sales Person 1': 78000.0,
    'Sales Person 2': 32000.0,
    'Sales Person 3': 45000.0,
    'Sales Person 4': 11000.0,
    'Sales Person 5': 20000.0
}

workbook = xlsxwriter.Workbook('sales_report.xlsx')         # Create a workbook and add a worksheet
worksheet = workbook.add_worksheet()

worksheet.write('A1', 'Sales Person')   # Add data to the worksheet
worksheet.write('B1', 'Sales Amount')

row = 1
for sales_person, amount in sales_data.items():
    worksheet.write(row, 0, sales_person)
    worksheet.write(row, 1, amount)
    row += 1

chart = workbook.add_chart({'type': 'pie'})     # Create a chart object

chart.add_series({                              # Configure the chart series
    'name': 'Sales Amount',
    'categories': [worksheet.name, 1, 0, len(sales_data), 0],
    'values': [worksheet.name, 1, 1, len(sales_data), 1],
})

chart.set_title({'name': 'Sales Distribution'})      # Set the chart title
worksheet.insert_chart('D2', chart)                 # Insert the chart into the worksheet
workbook.close()                                    # Close the workbook


'''Todo 4. Create a medals report in xlsx for olympics based on countries as shown below.
Country, Gold, Silver, Bronze, Total
Country 1, 15, 20, 25, 60
Country 2, 8,15,10, 33
Country 3, 12,25,10, 47
Country 4, 21,12,13, 50
Country 5, 18,29,30, 77
Create a Column chart with 3 medals gold, silver and bronze data.
Create another Line chart with only total medals data.'''

# import xlsxwriter
#
# medals= [
#     ['Country', 'Gold', 'Silver', 'Bronze', 'Total'],
#     ['Country 1', 15, 20, 25, 60],
#     ['Country 2', 8, 15, 10, 33],
#     ['Country 3', 12, 25, 10, 47],
#     ['Country 4', 21, 12, 13, 50],
#     ['Country 5', 18, 29, 30, 77]
# ]
#
# workbook = xlsxwriter.Workbook('medals_report.xlsx')
# worksheet = workbook.add_worksheet()
#
# for row_num, row_data in enumerate(medals):     # Add data to the worksheet
#     worksheet.write_row(row_num, 0, row_data)
#
# chart1 = workbook.add_chart({'type': 'column'})      # Create a column chart for Gold, Silver, and Bronze medals
#
# for col_num in range(1, 4):
#     chart1.add_series({
#         'name': [worksheet.name, 0, col_num],
#         'categories': [worksheet.name, 1, 0, 5, 0],
#         'values': [worksheet.name, 1, col_num, 5, col_num],
#     })
#
# chart1.set_title({'name': 'Medals by Country'})
# chart1.set_x_axis({'name': 'Country'})
# chart1.set_y_axis({'name': 'Medals'})
#
# worksheet.insert_chart('G2', chart1)
#
# chart2 = workbook.add_chart({'type': 'line'})       # Create a line chart for Total medals
#
# chart2.add_series({
#     'name': [worksheet.name, 0, 4],
#     'categories': [worksheet.name, 1, 0, 5, 0],
#     'values': [worksheet.name, 1, 4, 5, 4],
# })
#
# chart2.set_title({'name': 'Total Medals by Country'})
# chart2.set_x_axis({'name': 'Country'})
# chart2.set_y_axis({'name': 'Total Medals'})
#
# worksheet.insert_chart('G18', chart2)
#
# workbook.close()


'''Todo 5. Read an Excel file and prepare a data as shown below. Excel Data'''

import openpyxl

workbook = openpyxl.load_workbook('Exceldata.xlsx')     # Open the Excel file
sheet = workbook.active  # Get the first sheet

result_dict = {'Male': [], 'Female': []}    # Create an empty dictionary to store the data

for row in sheet.iter_rows(min_row=2, values_only=True):    # Iterate over each row in the worksheet
    name = row[0]
    surname = row[1]
    gender = row[2]
    age = int(row[3])           # Assuming age is an integer in the Excel file
    city = row[4]
    nationality = row[5]

    full_name = f"{name} {surname}"  # Combine name and surname

    row_dict = {'name': full_name.strip(), 'age': age, 'city': city, 'nationality': nationality}    # Create a dictionary for the row


    if gender == 'Male':                # Add the row dictionary to the appropriate list based on gender
        result_dict['Male'].append(row_dict)

    elif gender == 'Female':
        result_dict['Female'].append(row_dict)


print(result_dict)


'''Todo 6. Create a word document with header as Employee Experience Certificate, Write a paragraph for his experience description. Mention his 
projects as ordered list. Mention his skills with unordered list. Create a table with the following details'''

# from docx import Document
#
# doc = Document()
#
# header = doc.sections[0].header
# header_paragraph = header.paragraphs[0]
# header_paragraph.add_run("Employee Experience Certificate").bold=True
#
#
# experience_paragraph = doc.add_paragraph()
# experience_paragraph.add_run("Mahesh Patel has been an invaluable member of our team, bringing his expertise and dedication to"
#                              " every project undertaken. With 5 years of experience in IT industry, he has consistently delivered "
#                              "exceptional results and contributed significantly to our organization's success.")
#
# doc.add_paragraph("\nProject:")
# doc.add_paragraph(
#     'first project in ordered list', style='List Number'
# )
# doc.add_paragraph(
#     'second project in ordered list', style='List Number'
# )
# doc.add_paragraph(
#     'third project in ordered list', style='List Number'
# )
#
# # Add skills as unordered list
# skills = [
#     "Python programming",
#     "Data analysis",
#     "Problem-solving"
# ]
#
# doc.add_paragraph("\nSkills:")
# for skill in skills:
#     doc.add_paragraph(skill, style='List Bullet')
#
# # Add table
# table = doc.add_table(rows=2, cols=4)
# table.style = 'Table Grid'
#
# # Table Header
# hdr_cells = table.rows[0].cells
# hdr_cells[0].text = "Total no of years"
# hdr_cells[1].text = "Total Working Days"
# hdr_cells[2].text = "Worked Days"
# hdr_cells[3].text = "Leaves"
#
# # Table Data
# row_cells = table.rows[1].cells
# row_cells[0].text = "4"
# row_cells[1].text = "1056"
# row_cells[2].text = "1016"
# row_cells[3].text = "Total Leaves: 40\nPaid Leave: 25\nSick Leave: 5\nUnpaid Leave: 10"
#
# doc.save("Employee_Experience_Certificate.docx")


'''Todo 7. Read the data written in the above document file. For table data prepare a dictionary with their headers.'''


# from docx import Document
#
# doc = Document('Employee_Experience_Certificate.docx')  # Load the Word document
#
# table = {}
#
# for table in doc.tables:    # Read the table data
#     for i, row in enumerate(table.rows):
#
#         if i == 0:                      # Assuming the first row contains headers
#             headers = [cell.text.strip() for cell in row.cells]
#             table = {header: [] for header in headers}
#
#         else:
#             for j, cell in enumerate(row.cells):
#                 table[headers[j]].append(cell.text.strip())
#
# print(table)



'''Todo 8. Create a CSV from the following list of dictionary.
[
{‘name’:’Infosys’, ‘location’:’Pune’,’strength’:40,000.0},
{‘name’:’TCS’, ‘location’:’Gandhinagar’,’strength’:80,000.0},
{‘name’:’Wipro’, ‘location’:’Banguluru’,’strength’:65,000.0},
{‘name’:’Accenture’, ‘location’:’Gurugram’,’strength’:45,000.0},
{‘name’:’Capegemini’, ‘location’:’Mumbai’,’strength’:32,000.0},
]'''

# import csv
#
# data = [
#     {'name': 'Infosys', 'location': 'Pune', 'strength': 40000.0},
#     {'name': 'TCS', 'location': 'Gandhinagar', 'strength': 80000.0},
#     {'name': 'Wipro', 'location': 'Banguluru', 'strength': 65000.0},
#     {'name': 'Accenture', 'location': 'Gurugram', 'strength': 45000.0},
#     {'name': 'Capegemini', 'location': 'Mumbai', 'strength': 32000.0}
# ]
#
# csv_file = 'companies.csv'
#
# with open(csv_file, mode='w', newline='') as file:
#     fieldnames = ['name', 'location', 'strength']
#     writer = csv.DictWriter(file, fieldnames=fieldnames)
#
#     writer.writeheader()
#     for company in data:
#         writer.writerow(company)
#
# print("CSV file is Created Successfully!")



'''Todo 9. Read a CSV file and create a list of dictionary as shown below.CSV Data'''

import csv

def csv_to_dict(file):
    data = []

    with open(file, newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')
        next(reader)                # Skip the header row

        for row in reader:
            rp, establishment, users, cost = row
            data.append({
                'ERP': erp,
                'Establishment': int(establishment),
                'Users': int(users),
                'Cost': float(cost)
            })
    return data

file= 'ERP.csv'
dict_list = csv_to_dict(file)
for item in dict_list:
    print(item)
